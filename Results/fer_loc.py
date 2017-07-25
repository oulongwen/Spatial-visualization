from openpyxl import load_workbook

class ferplant:
    fertilizer = load_workbook('../Data/fertilizer plants.xlsx',
                     data_only=True, read_only = True)
    ferloc = fertilizer['Location']
    fer_dict = {}
    for row in ferloc.iter_rows('A3:D'+str(ferloc.max_row)):
        name = row[0].value
        if name != None:
            name = name.lower()
            if name not in fer_dict.keys():
                fer_dict[name] = (row[-2].value, row[-1].value)