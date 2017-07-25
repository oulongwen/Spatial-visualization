import codecs
from openpyxl import load_workbook

class fields:
    wb = load_workbook('../GIS emission/fcml.xlsx',data_only=True, read_only = True)
    ws = wb['2015 FCML']

    keys = []
    values = []

    for row in ws.iter_rows('A2:I66340'):
        try:
            key = str(row[0].value).strip() + str(row[1].value).strip()
            keys.append(key)
            values.append(row[-1].value)
        except:
            pass



    wb2 = load_workbook('../GIS emission/oil and gas fields.xlsx',data_only=True, read_only = True)
    ws2 = wb2['Oil fields']
    ws3 = wb2['Gas fields']

    oilfid=[]
    oilprod=[]

    occurrences = lambda s, lst: [i for i,e in enumerate(lst) if e == s]
    for row in ws2.iter_rows('B3:E106'):
        fieldname = str(row[0].value)
        try:
            index = fieldname.index('(')
            fieldname = fieldname[:index]
        except:
            pass
        fieldname = fieldname.strip()
        fieldname = fieldname + row[1].value.strip()
        occur = occurrences(fieldname, keys)
        if len(occur) > 0:
            for j in occur:
                oilfid.append(values[j])
                oilprod.append(row[-1].value/len(occur))
    
    gasfid=[]
    gasprod=[]
    
    for row in ws3.iter_rows('B2:E122'):
        fieldname = str(row[0].value)
        try:
            index = fieldname.index('(')
            fieldname = fieldname[:index]
        except:
            pass
        fieldname = fieldname.strip()
        fieldname = fieldname + row[1].value.strip()
        occur = occurrences(fieldname, keys)
        if len(occur) > 0:
            for j in occur:
                gasfid.append(values[j])
                gasprod.append(row[-1].value/len(occur))